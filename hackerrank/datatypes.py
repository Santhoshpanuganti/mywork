import numpy as np
from PIL import Image
import os
import cv2
import zipfile
import xlwt
from xlrd import open_workbook
import pandas as pd
from selenium import webdriver
from openpyxl import load_workbook
from email.mime.text import MIMEText
import smtplib
import socket
import xlsxwriter

def max_number():
    a = [1,4,8,10,90]
    ma  = 0
    # legnth = len(a)
    for i in a:
        if i > ma:
            ma = i
    print(ma)


def ispalindrome():
    string = 'sasa'
    string1 = string[::-1]
    if string1 == string:
        print(True)
    else:
        print(False)


def leap_year(year):
    leap = False
    if year % 4 == 0:
        if year % 100 == 0:
            if year % 400 == 0:
                return True
            else:
                return leap
        elif year % 400:
            return True
        else:
            return leap
    else:
        return leap



def get_avg_of_marks():
    n = int(input())
    student_marks = {}
    for _ in range(n):
        line = input().split()
        name, scores = line[0], line[1:]
        scores = map(float, scores)
        marks = [x for x in scores]
        total = 0
        for i in marks:
            total = total+i
        avg = total/len(marks)
        avg = format(avg,'.2f')
        student_marks[name] = scores,avg
    query_name = input()
    return student_marks.get(query_name,None)[1]


def on_numpy():
    n = int(input())
    a1 = []
    b1 = []
    for i in range(n):
        i = int(input())
        a1.append(i)
    for j in range(n):
        j = int(input())
        b1.append(j)
    a = np.array(a1)
    b = np.array(b1)
    outer = np.outer(a,b)
    inner = np.inner(a,b)
    return inner,outer


def runner_up_score():
    n = int(input())
    arr = map(int,input().split())
    a = [k for k in arr]
    b = a[0:n]
    for j in range(len(b)):
        for l in range(j+1,len(b)):
            if b[j] <= b[l]:
                temp = b[j]
                b[j] = b[l]
                b[l] = temp
    c = []
    for i in b:
        if i not in c:
            c.append(i)
    print(c)
    print(b)
    print(c[1])


def img_compare():
  images = []
  folder = 'D:/logos/jpeg/rolling_imgs'
  for file_name  in os.listdir(folder):
      img = cv2.imread(os.path.join(folder,file_name))
      if img is not None:
          images.append(img)
  return images


def image_attributes():
    folder = 'D:/logos/jpeg/rolling_imgs'
    folder_2 = 'D:/logos/jpeg/rolling_imgs - Copy'
    for file_name in os.listdir(folder):
        im = Image.open(os.path.join(folder,file_name),mode='r')
        width,height = im.size
        # print(file_name+'::'+'WIDTH'+'::'+str(width)+'::'+'HEIGHT'+'::'+str(height))
        for file_name_2 in os.listdir(folder_2):
            im2 = Image.open(os.path.join(folder_2,file_name_2),mode='r')
            width_2,height_2 =im2.size
            if width == width_2:
                print(True)
            else:
                print(file_name+'::'+str(width)+','+file_name_2+'::'+str(width_2))
            if height == height_2:
                print(True)
            else:
                print(file_name+'::'+str(height)+','+file_name_2+'::'+str(height_2))


def zip_files():
    zip = zipfile.ZipFile('D:/logos/pngs.zip',mode='w')
    zip.close()


def img_convert():
    paths = os.listdir('D:/logos/pngs/others')
    for file in paths:
        img = Image.open('D:/logos/pngs/others/'+file)
        rgb_img = img.convert('RGB')
        if '.png' in file:
            new_file = file.replace('.png','.jpeg')
            saved = rgb_img.save('D:/logos/jpeg/others'+new_file)


def a():
    wb = open_workbook('D:/exceldatascrapped/mobiles.xlsx')
    sheet = wb.sheet_by_index(1)
    for row in range(sheet.nrows):
        print(sheet.cell_value(row,0))
        print(sheet.cell_value(row,1))


def b():
    wb = xlwt.Workbook()
    sheet1 = wb.add_sheet('sheet1')
    for row in range(1000):
        sheet1.write(row,0,'title')
        sheet1.write(row,1,'price')
        sheet1.write(row,2,'specs')
        wb.save('D:/exceldatascrapped/sample.xlsx')


def panda():
    raws = {'price':[1245,1478],'title':['Samsung','Nokia']}
    df = pd.DataFrame(raws,columns=['price','title'])
    df.to_excel('D:/exceldatascrapped/sam.xlsx')


def click():
    browser = webdriver.Chrome(executable_path='D:/chromedriver/chromedriver.exe')
    browser.get('https://www.google.com/')
    ele = browser.find_element_by_link_text('https://accounts.google.com/ServiceLogin?hl=en&passive=true&continue=https://www.google.com/')
    browser.implicitly_wait(5)
    ele.click()


def to_existed_file():
    data = {'price':[1213,159],'product':['xyz','abc']}
    df1 = pd.DataFrame(data,columns=['price','product'])
    path = 'D:/exceldatascrapped/fashion/bata.xlsx'
    book = load_workbook(path)
    writer = pd.ExcelWriter(path,engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df1.to_excel(writer,sheet_name='fk1')
    writer.save()


def send_email():
    msg = MIMEText('Hello Santhosh')
    socket.getaddrinfo('smtp.mail.google.com',8080)
    s = smtplib.SMTP('smtp.mail.google.com')
    s.sendmail(msg['santoshsantu597@gmail.com'],msg['panuganti994@gmail.com'],msg.as_string())
    s.quit()


def create_excelfiles():
    file = os.listdir('D:/excelsheet_data')
    if 'amazonmobile.xlsx' not in file:
        sheet = xlsxwriter.Workbook('D:/excelsheet_data/amazonmobile.xlsx')
        sheet.add_worksheet()
    print(file)

def second_highest_score():
    t_info = []
    for i in range(int(input('enter count'))):
        name = input('enter name')
        score = float(input('enter score'))
        info = [name,score]
        t_info.append(info)
    print(t_info[0][1])
second_highest_score()